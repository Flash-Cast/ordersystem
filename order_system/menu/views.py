# Create your views here.
from django.shortcuts import render, redirect, get_object_or_404
from .models import MenuItem, Order, OrderItem
from django.views.decorators.csrf import csrf_exempt # 実運用ではCSRF保護を有効にしてください
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
import os
from django.conf import settings
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.db import transaction
import logging

# Excel関連の定数
EXCEL_FILE_NAME = 'orders.xlsx'
EXCEL_HEADERS = ["整理番号", "注文日時", "品目", "数量", "小計", "状態", "完了日時"]

logger = logging.getLogger(__name__)

def _get_excel_workbook(excel_path):
    if not os.path.exists(excel_path):
        return Workbook()
    try:
        return load_workbook(excel_path)
    except Exception as e:
        logger.error(f"Excelファイルの読み込みに失敗しました: {excel_path},エラー: {e}")
        return Workbook()

def _get_or_create_sheet_by_name(workbook, sheet_name): # headers引数を削除 (EXCEL_HEADERSを直接使用)
    if sheet_name not in workbook.sheetnames:
        ws = workbook.create_sheet(sheet_name)
        ws.append(EXCEL_HEADERS) # グローバル定数を使用
    else:
        ws = workbook[sheet_name]
    return ws

def menu_list(request):
    items = MenuItem.objects.all()
    return render(request, 'menu/menu_list.html', {'items': items})

def order_confirm(request):
    if request.method == 'POST':
        items = MenuItem.objects.all()
        order_items_data = []
        total = 0
        for item in items:
            quantity_str = request.POST.get(f'quantity_{item.id}', '0')
            try:
                qty = int(quantity_str)
                if qty < 0: qty = 0
            except ValueError:
                qty = 0
            if qty > 0:
                subtotal = item.price * qty
                order_items_data.append({
                    'id': item.id,
                    'name': item.name,
                    'price': item.price,
                    'quantity': qty,
                    'subtotal': subtotal
                })
                total += subtotal
        if not order_items_data:
            return redirect('menu_list')
        return render(request, 'menu/order_confirm.html', {
            'order_items': order_items_data,
            'total': total
        })
    return redirect('menu_list')

@csrf_exempt
@transaction.atomic
def order_submit(request):
    if request.method == 'POST':
        items = MenuItem.objects.all()
        order = Order.objects.create(total_price=0)
        current_total = 0
        order_created_items_for_excel = []
        for item_model in items:
            quantity_str = request.POST.get(f'quantity_{item_model.id}', '0')
            try:
                qty = int(quantity_str)
                if qty < 0: qty = 0
            except ValueError:
                qty = 0
            if qty > 0:
                OrderItem.objects.create(order=order, item=item_model, quantity=qty)
                current_total += item_model.price * qty
                order_created_items_for_excel.append({
                    'name': item_model.name,
                    'quantity': qty,
                    'subtotal': item_model.price * qty
                })
        if not order_created_items_for_excel:
             return HttpResponse("注文アイテムが選択されていません。 <a href=\"/\">戻る</a>")
        order.total_price = current_total
        order.save()

        excel_path = os.path.join(settings.BASE_DIR, EXCEL_FILE_NAME)
        wb = _get_excel_workbook(excel_path)
        
        # ★★★ 変更点: シート名を日ごとにする ★★★
        sheet_name_for_order = order.created_at.strftime("%Y-%m-%d") # 月ごとから日ごとに変更
        ws = _get_or_create_sheet_by_name(wb, sheet_name_for_order) # headers引数を削除

        order_time_str = order.created_at.strftime("%Y-%m-%d %H:%M:%S")
        for created_item in order_created_items_for_excel:
            ws.append([
                order.id,
                order_time_str,
                created_item['name'],
                created_item['quantity'],
                created_item['subtotal'],
                "処理中",
                ""
            ])
        try:
            wb.save(excel_path)
        except Exception as e:
            logger.error(f"Excelファイルの保存に失敗しました: {excel_path}, エラー: {e}")
        context = {
            'order_id': order.id,
            'total_price': order.total_price,
            'user': request.user
        }
        return render(request, 'menu/order_complete.html', context)
    return HttpResponse("無効なアクセスです。POSTメソッドで注文してください。 <a href=\"/\">戻る</a>")

@login_required
def order_list(request):
    orders = Order.objects.all().order_by('-created_at')
    return render(request, 'menu/order_list.html', {'orders': orders})

@login_required
@csrf_exempt
@transaction.atomic
def complete_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if order.completed_at:
        return redirect('order_list')

    # ★★★ 変更点: 完了情報を書き込むシートも、元の注文が作成された「日」のシートにする ★★★
    original_order_created_day_sheet_name = order.created_at.strftime("%Y-%m-%d") # 月ごとから日ごとに変更
    order.completed_at = timezone.now()
    order.save()

    excel_path = os.path.join(settings.BASE_DIR, EXCEL_FILE_NAME)
    wb = _get_excel_workbook(excel_path)
    
    ws = _get_or_create_sheet_by_name(wb, original_order_created_day_sheet_name) # headers引数を削除
    
    ws.append([
        order.id,
        order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        "【注文完了イベント】",
        "",
        order.total_price,
        "完了",
        order.completed_at.strftime("%Y-%m-%d %H:%M:%S")
    ])
    try:
        wb.save(excel_path)
    except Exception as e:
        logger.error(f"Excelファイルの保存中にエラーが発生しました (注文完了時): {excel_path}, エラー: {e}")
    return redirect('order_list')